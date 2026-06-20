#!/usr/bin/env python3
"""Generate chapter infographics for chapters 5-8 of 对比Excel，轻松学习Python报表自动化"""
import os

BOOKDIR = "/Users/jessper/.openclaw/workspace/infographics/books"
BASE = "对比Excel，轻松学习Python报表自动化"
FONT = "../方正屏显雅宋简体.TTF"

CSS = '''
  @font-face {font-family:'FZXPYZS';src:url('%s') format('truetype');font-weight:normal;font-style:normal}
  *{margin:0;padding:0;box-sizing:border-box}
  body{background:#f5f1eb;font-family:'PingFang SC','Noto Serif SC','STSong',Georgia,serif;display:flex;justify-content:center;align-items:flex-start;min-height:100vh;padding:40px 20px 60px}
  .container{max-width:880px;width:100%}
  h1{font-family:'FZXPYZS','PingFang SC','Noto Serif SC',serif;font-size:36px;color:#1a1a1a;text-align:center;line-height:1.4;margin-bottom:8px;font-weight:normal;letter-spacing:1.5px}
  .subtitle{text-align:center;font-family:'FZXPYZS','PingFang SC','Noto Serif SC',serif;font-size:14px;color:#888;margin-bottom:24px;line-height:1.7;max-width:640px;margin-left:auto;margin-right:auto}
  .divider{width:60px;height:3px;background:linear-gradient(90deg,#dc2626,#ea580c);margin:0 auto 28px;border-radius:2px}
  .chapter-overview{background:#f8f6f3;border-left:3px solid #4f46e5;border-radius:8px;padding:16px 20px;margin:12px 0 24px;font-size:14px;color:#555;line-height:1.8;font-family:'FZXPYZS','PingFang SC',serif}
  .chapter-overview p{margin:0}
  .kpi-row{display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:12px;margin-bottom:18px}
  .kpi-card{background:#ffffff;border:1px solid #e8e0d5;border-radius:14px;padding:20px 14px;text-align:center;box-shadow:0 1px 3px rgba(0,0,0,0.04)}
  .kpi-num{font-size:28px;font-weight:bold;color:#1a1a1a;margin-bottom:4px;font-family:'FZXPYZS','PingFang SC',serif}
  .kpi-label{font-size:12px;color:#888;line-height:1.5;font-family:'FZXPYZS','PingFang SC',serif}
  .section{background:#ffffff;border-radius:14px;margin-bottom:18px;padding:24px 28px;box-shadow:0 1px 3px rgba(0,0,0,0.04);display:flex;gap:20px;align-items:flex-start;border-left:4px solid transparent}
  .section-01{border-left-color:#dc2626}.section-02{border-left-color:#ea580c}.section-03{border-left-color:#ca8a04}.section-04{border-left-color:#4f46e5}.section-05{border-left-color:#db2777}
  .section-num{flex-shrink:0;width:42px;height:42px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:17px;font-weight:bold;margin-top:2px;font-family:'FZXPYZS','PingFang SC',serif}
  .num-01{background:#fef2f2;color:#dc2626}.num-02{background:#fff7ed;color:#ea580c}.num-03{background:#fefce8;color:#ca8a04}.num-04{background:#eef2ff;color:#4f46e5}.num-05{background:#fdf2f8;color:#db2777}
  .section-body{flex:1}
  .tag{display:inline-block;font-size:11px;font-weight:bold;padding:2px 10px;border-radius:10px;margin-bottom:8px;letter-spacing:1px;font-family:'FZXPYZS','PingFang SC',serif}
  .tag-01{background:#fef2f2;color:#dc2626}.tag-02{background:#fff7ed;color:#ea580c}.tag-03{background:#fefce8;color:#ca8a04}.tag-04{background:#eef2ff;color:#4f46e5}.tag-05{background:#fdf2f8;color:#db2777}
  .section-title{font-size:18px;margin-bottom:10px;font-weight:bold;line-height:1.4;font-family:'FZXPYZS','PingFang SC',serif}
  .t-01{color:#dc2626}.t-02{color:#ea580c}.t-03{color:#ca8a04}.t-04{color:#4f46e5}.t-05{color:#db2777}
  .section-desc{font-size:14px;color:#555;line-height:1.9;margin-bottom:14px}
  .flow-row{display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-top:6px}
  .flow-step{background:#fff7ed;border:1px solid #fed7aa;border-radius:10px;padding:10px 12px;text-align:center;min-width:80px;flex:1;font-size:13px;color:#9a3412;line-height:1.5;font-weight:bold}
  .flow-arrow{font-size:20px;color:#ea580c;flex-shrink:0;font-weight:bold}
  .flow-step.end{background:#fef2f2;border-color:#fecaca;color:#991b1b}
  .dual-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-top:10px}
  .dual-card{border-radius:12px;padding:18px 20px;display:flex;gap:12px;align-items:flex-start}
  .dual-card.yes{background:#f0fdf4;border:1px solid #bbf7d0}
  .dual-card.no{background:#fef2f2;border:1px solid #fecaca}
  .dual-icon{font-size:24px;flex-shrink:0;line-height:1}
  .dual-text h4{font-size:14px;color:#1a1a1a;margin-bottom:4px;font-family:'FZXPYZS','PingFang SC',serif}
  .dual-text p{font-size:12px;color:#777;line-height:1.6}
  .cmp-grid{display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin-top:10px}
  .cmp-card{border-radius:12px;padding:18px 14px;text-align:center;border:1px solid #e8e0d5}
  .cmp-card.otc{background:#f0fdf4;border-color:#bbf7d0}.cmp-card.etf{background:#fff7ed;border-color:#fed7aa}.cmp-card.lof{background:#eef2ff;border-color:#c7d2fe}
  .cmp-icon{font-size:28px;margin-bottom:6px}
  .cmp-name{font-size:15px;font-weight:bold;color:#1a1a1a;margin-bottom:4px;font-family:'FZXPYZS','PingFang SC',serif}
  .cmp-note{font-size:12px;color:#666;line-height:1.6}
  .code-block{background:#f4f0ea;border:1px solid #e8e0d5;border-radius:8px;padding:14px 18px;margin-top:10px;font-family:'SF Mono','Menlo','Consolas',monospace;font-size:12px;color:#555;line-height:1.7;overflow-x:auto;white-space:pre-wrap}
  .data-row{display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin-top:10px}
  .data-card{border-radius:12px;padding:20px 16px;text-align:center;border:1px solid #fce7f3;background:#fdf2f8}
  .data-big{font-size:28px;margin-bottom:6px}
  .data-name{font-size:15px;font-weight:bold;color:#831843;margin-bottom:6px;font-family:'FZXPYZS','PingFang SC',serif}
  .data-sm{font-size:12px;color:#9d174d;line-height:1.6}
  .takeaway{background:#ffffff;border:1px solid #e8e0d5;border-radius:14px;padding:24px 32px;margin-bottom:18px;box-shadow:0 1px 3px rgba(0,0,0,0.04);border-left:4px solid #dc2626}
  .takeaway-label{font-family:'FZXPYZS','PingFang SC',serif;font-size:12px;color:#dc2626;letter-spacing:2px;margin-bottom:6px;font-weight:bold}
  .takeaway-text{font-size:16px;color:#1a1a1a;line-height:1.9;font-family:'FZXPYZS','PingFang SC',serif}
  .footer{text-align:center;margin-top:32px;padding-top:20px;border-top:1px solid #e8e0d5;color:#bbb;font-size:13px;line-height:1.8}
  .lang-switch{text-align:right;margin-bottom:16px}
  .lang-btn{display:inline-block;padding:6px 16px;border-radius:8px;font-size:13px;text-decoration:none;letter-spacing:.03em;font-family:'FZXPYZS','PingFang SC',serif;background:#fef2f2;color:#dc2626;border:1px solid #fecaca;transition:opacity .15s}
  .lang-btn:hover{opacity:.75}
  .back-catalog{text-align:right;margin-bottom:4px}
  .back-catalog-btn{display:inline-block;padding:5px 14px;border-radius:8px;font-size:12px;text-decoration:none;letter-spacing:.02em;background:#eef2ff;color:#4f46e5;border:1px solid #c7d2fe;transition:opacity .15s}
  .back-catalog-btn:hover{opacity:.75}
  @media(max-width:640px){.section{flex-direction:column;align-items:center;text-align:center;border-left:none;border-top:4px solid transparent;padding-top:20px}.section-01{border-top-color:#dc2626}.section-02{border-top-color:#ea580c}.section-03{border-top-color:#ca8a04}.section-04{border-top-color:#4f46e5}.section-05{border-top-color:#db2777}.dual-grid,.data-row,.kpi-row,.cmp-grid{grid-template-columns:1fr}.flow-row{flex-direction:column}.flow-arrow{transform:rotate(90deg)}.container{padding:0 8px}h1{font-size:26px}}
''' % FONT


def build_html(ch_num, zh_data):
    """Build ZH HTML for a chapter"""
    ch_pad = f"ch{ch_num:03d}"
    ch_name = zh_data['ch_name']
    scqa = zh_data['scqa']
    overview = zh_data['overview']
    kpis_html = zh_data['kpis_html']
    sections_html = zh_data['sections_html']
    takeaway = zh_data['takeaway']
    footer_ch = zh_data['footer_ch']
    catalog = f"{BASE}-catalog.html"
    en_link = f"{BASE}-{ch_pad}-info-en.html"

    return f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{BASE} · 第{ch_num}章「{ch_name}」</title>
<style>
{CSS}
</style>
</head>
<body>
<div class="container">
<div class="back-catalog"><a class="back-catalog-btn" href="{catalog}">← 返回章节目录</a></div>
<div class="lang-switch">
  <a class="lang-btn" target="_blank" href="{en_link}">中文 / English</a>
</div>

<h1>{BASE} · 第{ch_num}章「{ch_name}」</h1>
<p class="subtitle">
  <span class="scqa-label" style="display:block;margin-bottom:6px;">SCQA · 情境 → 冲突 → 问题 → 回答</span>
  {scqa}
</p>
<div class="divider"></div>
<div class="chapter-overview">
  <p>{overview}</p>
</div>

{kpis_html}
{sections_html}

<div class="takeaway">
  <div class="takeaway-label">🔑 核心结论</div>
  <div class="takeaway-text">{takeaway}</div>
</div>

<div class="footer">
  来源：张俊红《{BASE}》第{ch_num}章「{footer_ch}」· 信息图 · 仅供学习参考
</div>

</div>
</body>
</html>'''


# ===== Chapter 5 =====
zh5 = {
    'ch_name': '用Python设置Excel对齐方式',
    'scqa': '在Excel中调整对齐方式——左对齐、居中、右对齐、自动换行、合并单元格——只需在「开始」选项卡的「对齐方式」组点击几次按钮。但当你有数百行数据需要统一对齐、或每天生成的报表都要标准化排版时，手工调整就枯燥低效。如何用Python批量控制Excel单元格的对齐方式？本章系统讲解Alignment()函数的6大参数、水平/垂直对齐类型、合并/解除单元格、以及批量对齐的实战案例。',
    'overview': '本章聚焦于Excel「开始」选项卡中「对齐方式」组的Python实现。首先介绍Alignment()函数的6个核心参数：horizontal（水平对齐，支持left/center/right/justify/distributed/fill等9种）、vertical（垂直对齐，top/center/bottom）、text_rotation（文本旋转角度，0-360°）、wrap_text（自动换行）、shrink_to_fit（自适应缩小字号）、indent（缩进字符数）。然后讲解单元格合并/解除合并（merge_cells/unmerge_cells），以及合并后单元格通过top_left_cell进行样式设置的方法。最后通过一个批量设置对齐方式加合并标题行的综合案例，展示如何用for循环一次性完成整表标准化排版。',
    'kpis_html': '''
<div class="kpi-row">
  <div class="kpi-card">
    <div class="kpi-num">6个</div>
    <div class="kpi-label">Alignment()参数<br>水平/垂直/旋转等</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-num">9种</div>
    <div class="kpi-label">水平对齐类型<br>居左/居中/居右等</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-num">3种</div>
    <div class="kpi-label">垂直对齐类型<br>顶部/居中/底部</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-num">∞</div>
    <div class="kpi-label">批量对齐<br>遍历整表统一排版</div>
  </div>
</div>''',
    'sections_html': '''
<div class="section section-01">
  <div class="section-num num-01">01</div>
  <div class="section-body">
    <div class="tag tag-01">核心函数</div>
    <div class="section-title t-01">Alignment()：六参数掌控对齐全局</div>
    <div class="section-desc">Alignment(horizontal, vertical, text_rotation, wrap_text, shrink_to_fit, indent) 是 openpyxl 对齐设置的核心。设置方式：cell.alignment = Alignment(参数1=值1, ...)。6个参数直接映射Excel功能区按钮，学习曲线平缓。</div>
    <div class="cmp-grid">
      <div class="cmp-card otc">
        <div class="cmp-icon">↔️</div>
        <div class="cmp-name">horizontal</div>
        <div class="cmp-note">水平对齐<br>left/center/right<br>justify/distributed/fill</div>
      </div>
      <div class="cmp-card etf">
        <div class="cmp-icon">↕️</div>
        <div class="cmp-name">vertical</div>
        <div class="cmp-note">垂直对齐<br>top/center/bottom<br>上下方向控制</div>
      </div>
      <div class="cmp-card lof">
        <div class="cmp-icon">🔄</div>
        <div class="cmp-name">旋转·换行·缩进</div>
        <div class="cmp-note">text_rotation (0-360°)<br>wrap_text (布尔值)<br>indent (缩进字符数)</div>
      </div>
    </div>
  </div>
</div>

<div class="section section-02">
  <div class="section-num num-02">02</div>
  <div class="section-body">
    <div class="tag tag-02">水平对齐</div>
    <div class="section-title t-02">水平方向对齐：9种类型从常规到特殊</div>
    <div class="section-desc">horizontal参数支持9种对齐类型：left（左对齐）、center（居中）、right（右对齐）、justify（两端对齐，占满行靠两边界，未满行左对齐）、distributed（分散对齐，未满行自动调字间距占满）、fill（填满对齐，重复字符填充）、general（一般对齐）、centerContinuous、distributed。日常最常用的是left/center/right三种。</div>
    <div class="flow-row">
      <div class="flow-step">⚫ 左对齐<br><small>left</small></div>
      <div class="flow-arrow">→</div>
      <div class="flow-step">⚫ 居中对齐<br><small>center</small></div>
      <div class="flow-arrow">→</div>
      <div class="flow-step">⚫ 右对齐<br><small>right</small></div>
      <div class="flow-arrow">→</div>
      <div class="flow-step">📐 两端对齐<br><small>justify</small></div>
      <div class="flow-arrow">→</div>
      <div class="flow-step end">📏 分散对齐<br><small>distributed</small></div>
    </div>
  </div>
</div>

<div class="section section-03">
  <div class="section-num num-03">03</div>
  <div class="section-body">
    <div class="tag tag-03">换行与旋转</div>
    <div class="section-title t-03">自动换行、文本旋转与字号自适应</div>
    <div class="section-desc">当文本超过列宽时三种处理：wrap_text=True自动换行，文本在单元格内分行显示；shrink_to_fit=True自动缩小字体适配单元格；text_rotation设置旋转角度（逆时针）。text_rotation与shrink_to_fit不能同时使用，各有适用场景——换行适合备注列，自适应适合紧凑表格，旋转适合窄列表头。</div>
    <div class="dual-grid">
      <div class="dual-card yes">
        <div class="dual-icon">📝</div>
        <div class="dual-text"><h4>自动换行（wrap_text=True）</h4><p>超列宽自动换行<br>字体大小不变<br>行高自动增加<br>适合长文本列</p></div>
      </div>
      <div class="dual-card no">
        <div class="dual-icon">🔎</div>
        <div class="dual-text"><h4>自适应+旋转</h4><p>shrink_to_fit 缩小字体<br>text_rotation 旋转角度<br>indent 缩进字符<br>各有独立适用场景</p></div>
      </div>
    </div>
  </div>
</div>

<div class="section section-04">
  <div class="section-num num-04">04</div>
  <div class="section-body">
    <div class="tag tag-04">合并操作</div>
    <div class="section-title t-04">单元格合并与解除：merge_cells / unmerge_cells</div>
    <div class="section-desc">merge_cells('A2:D2')合并指定区域，unmerge_cells('A2:D2')解除合并。合并后单元格以左上角单元格（top_left_cell）为代表进行样式设置——字体、填充、对齐方式均作用于此。这一设计保证了合并区域样式的一致性。</div>
    <div class="code-block">from openpyxl.styles import Alignment, Font

# 合并单元格
ws.merge_cells('A2:D2')

# 通过左上角单元格设置样式
top_left_cell = ws['A2']
top_left_cell.alignment = Alignment(
    horizontal='center', vertical='center')
top_left_cell.font = Font(bold=True)

# 解除合并
ws.unmerge_cells('A2:D2')</div>
  </div>
</div>

<div class="section section-05">
  <div class="section-num num-05">05</div>
  <div class="section-body">
    <div class="tag tag-05">批量实战</div>
    <div class="section-title t-05">批量设置对齐：整表标准化一个循环搞定</div>
    <div class="section-desc">真实场景极少只格式化一个单元格。在字体格式化的for循环基础上增加alignment属性，即可一次性完成水平居中、垂直居中、换行设置。配合标题行合并与加粗，代码编写一次，永久适用于同类报表。</div>
    <div class="data-row">
      <div class="data-card">
        <div class="data-big">📐</div>
        <div class="data-name">水平居中</div>
        <div class="data-sm">遍历所有数据单元格<br>horizontal='center'<br>整列数据统一居中</div>
      </div>
      <div class="data-card">
        <div class="data-big">📏</div>
        <div class="data-name">垂直居中</div>
        <div class="data-sm">遍历所有数据单元格<br>vertical='center'<br>行高统一视觉美观</div>
      </div>
      <div class="data-card">
        <div class="data-big">🏷️</div>
        <div class="data-name">标题行</div>
        <div class="data-sm">merge_cells合并<br>Font(bold=True)加粗<br>一次编写永久复用</div>
      </div>
    </div>
  </div>
</div>''',
    'takeaway': 'Excel的对齐方式——水平对齐、垂直对齐、自动换行、文本旋转、缩进、合并单元格——在Python中对应一个核心函数Alignment()和两个操作函数merge_cells()/unmerge_cells()。参数直接映射Excel功能区按钮，学习成本极低。真正的威力在于批量：一个for循环遍历所有数据单元格，统一设置alignment属性，配合标题行合并加粗，让报表排版和数据处理一样全自动化。这就是Python报表自动化的核心价值——取代一切重复、枯燥的工作。',
    'footer_ch': '用Python设置Excel对齐方式'
}

html5 = build_html(5, zh5)
path5 = os.path.join(BOOKDIR, f"{BASE}-ch005-info-zh.html")
with open(path5, 'w', encoding='utf-8') as f:
    f.write(html5)
print(f"✅ ch005 ZH: {len(html5)} bytes -> {path5}")


# ===== Chapter 6 =====
zh6 = {
    'ch_name': '用Python设置Excel数字、条件格式',
    'scqa': 'Excel的数字格式和条件格式是数据可视化的重要工具——将0.5显示为50%、将日期显示为年月日、将超标的单元格标红、给数据加上色阶或数据条。手工操作时，你要逐个选中区域、点击条件格式菜单、设置规则。但当报表每天更新且数据量庞大时，重复做这些事令人崩溃。如何用Python自动设置Excel的数字显示格式和条件格式？本章详细讲解number_format的9种数字格式、CellIsRule突出显示、DataBarRule数据条、ColorScaleRule色阶、IconSet图标集。',
    'overview': '本章分为两大板块：数字格式和条件格式。数字格式部分，number_format属性对应Excel「数字」组，覆盖常规、小数（0.00）、百分比（0.00%）、货币（¥#,##0）、科学计数、日期时间（yyyy-mm-dd h:mm:ss）等9种常用格式。条件格式部分包括四种类型：CellIsRule()用于突出显示满足条件的单元格（支持大于/小于/介于/等于等operator）；DataBarRule()用于数据条可视化（可自定义min/max类型、颜色、是否显示数值）；ColorScaleRule()用于色阶（双色和三色刻度）；IconSet()用于图标集（多类图标可选、可设置反转和是否显示数值）。两部分共同构成Excel数据可视化的Python自动化方案。',
    'kpis_html': '''
<div class="kpi-row">
  <div class="kpi-card">
    <div class="kpi-num">9种</div>
    <div class="kpi-label">数字格式<br>常规/小数/百分比等</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-num">4类</div>
    <div class="kpi-label">条件格式<br>高亮/数据条/色阶/图标</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-num">7种</div>
    <div class="kpi-label">operator条件<br>大于/小于/介于等</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-num">5种</div>
    <div class="kpi-label">图标集类型<br>三色旗/箭头/信号等</div>
  </div>
</div>''',
    'sections_html': '''
<div class="section section-01">
  <div class="section-num num-01">01</div>
  <div class="section-body">
    <div class="tag tag-01">数字格式</div>
    <div class="section-title t-01">number_format：9种数字显示格式一键切换</div>
    <div class="section-desc">number_format属性直接映射Excel「数字」组的功能区。常用格式：常规、0.00（保留2位小数）、0.00%（百分比）、¥#,##0（货币）、0.00E+00（科学计数）、yyyy-mm-dd h:mm:ss（日期时间）。设置方式：cell.number_format = '0.00'，一行代码替代Excel的手动点击。</div>
    <div class="flow-row">
      <div class="flow-step">🔢 常规<br><small>General</small></div>
      <div class="flow-arrow">→</div>
      <div class="flow-step">📊 小数<br><small>0.00</small></div>
      <div class="flow-arrow">→</div>
      <div class="flow-step">% 百分比<br><small>0.00%</small></div>
      <div class="flow-arrow">→</div>
      <div class="flow-step">¥ 货币<br><small>¥#,##0</small></div>
      <div class="flow-arrow">→</div>
      <div class="flow-step end">📅 日期<br><small>yyyy-mm-dd</small></div>
    </div>
  </div>
</div>

<div class="section section-02">
  <div class="section-num num-02">02</div>
  <div class="section-body">
    <div class="tag tag-02">突出显示</div>
    <div class="section-title t-02">CellIsRule：按条件自动标记异常数据</div>
    <div class="section-desc">conditional_formatting + CellIsRule(operator, formula, fill) 实现标准条件格式。operator支持greaterThan、lessThan、between、equal等7种。formula指定阈值（between用列表），fill定义满足条件时的填充颜色。常用于标记超过预算的支出、低于目标的KPI、区间内的合格数据。</div>
    <div class="code-block">from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

# 大于50的单元格填充红色
ws.conditional_formatting.add('A1:A10',
    CellIsRule(operator='greaterThan', formula=['50'],
               fill=PatternFill(fill_type='solid', fgColor='FF0000')))</div>
  </div>
</div>

<div class="section section-03">
  <div class="section-num num-03">03</div>
  <div class="section-body">
    <div class="tag tag-03">数据条</div>
    <div class="section-title t-03">DataBarRule：让数据大小一目了然</div>
    <div class="section-desc">DataBarRule(start_type, start_value, end_type, end_value, color, showValue) 将数值以数据条形式可视化。可自定义min/max的类型（min/max/num/percent/percentile/formula）和值、数据条颜色、是否显示数值。同时应用于多列只需改区域范围如'A1:B10'。</div>
    <div class="dual-grid">
      <div class="dual-card yes">
        <div class="dual-icon">📊</div>
        <div class="dual-text"><h4>默认数据条</h4><p>start_type='min'<br>end_type='max'<br>自动取最小/最大值<br>显示数值+数据条</p></div>
      </div>
      <div class="dual-card no">
        <div class="dual-icon">🎨</div>
        <div class="dual-text"><h4>自定义数据条</h4><p>start_type='num', value=0<br>end_type='num', value=100<br>自定义颜色+隐藏数值</p></div>
      </div>
    </div>
  </div>
</div>

<div class="section section-04">
  <div class="section-num num-04">04</div>
  <div class="section-body">
    <div class="tag tag-04">色阶</div>
    <div class="section-title t-04">ColorScaleRule：双色和三色渐变直观对比</div>
    <div class="section-desc">ColorScaleRule(start_type, start_value, start_color, mid_type..., end_type...) 用颜色渐变展示数值大小。双色刻度只需start+end参数，三色刻度增加mid中间值。常用于热力图、成绩分布、温度数据的可视化，替代手工调色。</div>
    <div class="cmp-grid">
      <div class="cmp-card otc">
        <div class="cmp-icon">🔴🟢</div>
        <div class="cmp-name">双色刻度</div>
        <div class="cmp-note">最小值 → 最大值<br>两种颜色渐变<br>红→绿/蓝→白等</div>
      </div>
      <div class="cmp-card etf">
        <div class="cmp-icon">🔴🟡🟢</div>
        <div class="cmp-name">三色刻度</div>
        <div class="cmp-note">最小值 → 中间 → 最大值<br>三色渐变过渡<br>红→黄→绿等</div>
      </div>
      <div class="cmp-card lof">
        <div class="cmp-icon">⚙️</div>
        <div class="cmp-name">自定义</div>
        <div class="cmp-note">可设min/mid/max类型<br>可自定义颜色<br>可设具体阈值</div>
      </div>
    </div>
  </div>
</div>

<div class="section section-05">
  <div class="section-num num-05">05</div>
  <div class="section-body">
    <div class="tag tag-05">图标集</div>
    <div class="section-title t-05">IconSet：用图标直观展示数据等级</div>
    <div class="section-desc">IconSet(iconSet, percent, cfvo, showValue, reverse) 将数据映射为图标。iconSet指定图标类型（3Arrows/3Flags/3TrafficLights等），cfvo列表定义各等级阈值，showValue控制是否显示数值，reverse控制是否反转图标方向。如同手机信号格，让数据等级一目了然。</div>
    <div class="data-row">
      <div class="data-card">
        <div class="data-big">🏁</div>
        <div class="data-name">三色旗</div>
        <div class="data-sm">3Flags图标<br>分三个等级<br>展示具体数值</div>
      </div>
      <div class="data-card">
        <div class="data-big">⬆️➡️⬇️</div>
        <div class="data-name">三向箭头</div>
        <div class="data-sm">3Arrows图标<br>分三个等级<br>可隐藏数值</div>
      </div>
      <div class="data-card">
        <div class="data-big">🚦</div>
        <div class="data-name">红绿灯</div>
        <div class="data-sm">3TrafficLights<br>支持4-5图标<br>可反转方向</div>
      </div>
    </div>
  </div>
</div>''',
    'takeaway': 'Excel的数字格式和条件格式是数据呈现的利器。Python中number_format属性覆盖常规、小数、百分比、货币、日期等9种常用数字格式。条件格式通过四种规则函数——CellIsRule（突出显示）、DataBarRule（数据条）、ColorScaleRule（色阶）、IconSet（图标集）——完整复现Excel条件格式的全部功能。它们的参数直接映射Excel对话框中的选项，学习路径清晰。批量应用只需将区域范围从'A1:A10'改为'A1:C20'，同一规则即可作用于多列数据。数字格式化与条件可视化全部自动化，报表的可读性和专业性大幅提升。',
    'footer_ch': '用Python设置Excel数字、条件格式'
}

html6 = build_html(6, zh6)
path6 = os.path.join(BOOKDIR, f"{BASE}-ch006-info-zh.html")
with open(path6, 'w', encoding='utf-8') as f:
    f.write(html6)
print(f"✅ ch006 ZH: {len(html6)} bytes -> {path6}")


# ===== Chapter 7 =====
zh7 = {
    'ch_name': '用Python设置Excel单元格',
    'scqa': 'Excel工作表的日常维护离不开插入行、删除行、调整行高列宽、隐藏区域。手工操作时，右键→插入→右键→删除→右键→行高→右键→列宽——重复无数次。当报表有几十个Sheet且每个都需要相同调整时，手工操作不仅耗时，还容易错删误改。如何用Python自动化管理Excel的单元格行列结构？本章讲解insert_rows/insert_cols、delete_rows/delete_cols、行高列宽设置、行列隐藏，以及批量遍历设置整表的实战案例。',
    'overview': '本章聚焦于Excel「开始」选项卡「单元格」组的Python实现，涵盖四大操作：①插入行/列——insert_rows(m, n)和insert_cols(m, n)，m为位置，n为数量（默认1）；②删除行/列——delete_rows(m, n)和delete_cols(m, n)，参数含义相同，是插入的逆向操作；③行高/列宽设置——通过row_dimensions[row_num].height和column_dimensions[col_letter].width分别设置；④行列隐藏——通过column_dimensions.group和row_dimensions.group创建组合来隐藏（Python中无直接隐藏API）。最后通过一个批量案例展示如何遍历多行多列统一设置行高列宽，实现整表格式标准化。',
    'kpis_html': '''
<div class="kpi-row">
  <div class="kpi-card">
    <div class="kpi-num">2+2</div>
    <div class="kpi-label">增删函数<br>insert & delete<br>行/列各一对</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-num">2种</div>
    <div class="kpi-label">尺寸属性<br>row_dimensions<br>column_dimensions</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-num">组合</div>
    <div class="kpi-label">隐藏方式<br>group创建组合<br>可折叠展开</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-num">∞</div>
    <div class="kpi-label">批量遍历<br>整表多行多列<br>一次循环搞定</div>
  </div>
</div>''',
    'sections_html': '''
<div class="section section-01">
  <div class="section-num num-01">01</div>
  <div class="section-body">
    <div class="tag tag-01">插入操作</div>
    <div class="section-title t-01">insert_rows / insert_cols：动态扩展工作表结构</div>
    <div class="section-desc">insert_rows(m, n)在第m行前插入n行，insert_cols(m, n)在第m列前插入n列。n默认值为1，可省略。这是Excel右键「插入」功能的Python等价操作，适用于动态添加数据区域、预留空行、插入表头等场景。</div>
    <div class="flow-row">
      <div class="flow-step">📍 定位<br><small>指定m位置</small></div>
      <div class="flow-arrow">→</div>
      <div class="flow-step">📥 插入行<br><small>insert_rows(m,n)</small></div>
      <div class="flow-arrow">→</div>
      <div class="flow-step">📊 插入列<br><small>insert_cols(m,n)</small></div>
      <div class="flow-arrow">→</div>
      <div class="flow-step">📋 数据下移<br><small>原数据自动移位</small></div>
      <div class="flow-arrow">→</div>
      <div class="flow-step end">✅ 完成<br><small>可复制Sheet对比</small></div>
    </div>
  </div>
</div>

<div class="section section-02">
  <div class="section-num num-02">02</div>
  <div class="section-body">
    <div class="tag tag-02">删除操作</div>
    <div class="section-title t-02">delete_rows / delete_cols：精准删除多余行列</div>
    <div class="section-desc">delete_rows(m, n)从第m行开始删除n行，delete_cols(m, n)从第m列开始删除n列。n默认值为1。这是插入的逆向操作，用于清理空行、删除不需要的数据列、精简报表结构。小心：删除后数据不可恢复。</div>
    <div class="dual-grid">
      <div class="dual-card yes">
        <div class="dual-icon">📥</div>
        <div class="dual-text"><h4>插入（insert）</h4><p>insert_rows(m, n)<br>insert_cols(m, n)<br>在指定位置前插入<br>n默认=1可不写</p></div>
      </div>
      <div class="dual-card no">
        <div class="dual-icon">🗑️</div>
        <div class="dual-text"><h4>删除（delete）</h4><p>delete_rows(m, n)<br>delete_cols(m, n)<br>从指定位置起删除<br>操作不可逆</p></div>
      </div>
    </div>
  </div>
</div>

<div class="section section-03">
  <div class="section-num num-03">03</div>
  <div class="section-body">
    <div class="tag tag-03">尺寸设置</div>
    <div class="section-title t-03">行高与列宽：row_dimensions / column_dimensions</div>
    <div class="section-desc">行高通过ws.row_dimensions[row_num].height设置，列宽通过ws.column_dimensions[col_letter].width设置。这与Excel右键「行高」「列宽」对话框完成相同的功能。改其他行只需换行号数字，改其他列只需换列字母。</div>
    <div class="code-block"># 设置第1行行高为40，第A列列宽为20
ws.row_dimensions[1].height = 40
ws.column_dimensions['A'].width = 20

# 改行号或列字母即可调整其他行列
ws.row_dimensions[2].height = 30
ws.column_dimensions['B'].width = 25</div>
  </div>
</div>

<div class="section section-04">
  <div class="section-num num-04">04</div>
  <div class="section-body">
    <div class="tag tag-04">隐藏操作</div>
    <div class="section-title t-04">行列隐藏：通过组合（group）实现可折叠区域</div>
    <div class="section-desc">openpyxl中隐藏行列的方式是创建组合（group）。column_dimensions.group(start, end, hidden=True)隐藏列，row_dimensions.group隐藏行。这与Excel右键「隐藏」稍有不同——group方式支持折叠/展开，更灵活。修改起止参数即可隐藏任意区域。</div>
    <div class="cmp-grid">
      <div class="cmp-card otc">
        <div class="cmp-icon">📁</div>
        <div class="cmp-name">隐藏行</div>
        <div class="cmp-note">ws.row_dimensions.group(7, 10)<br>隐藏第7-10行<br>修改起止即可</div>
      </div>
      <div class="cmp-card etf">
        <div class="cmp-icon">📂</div>
        <div class="cmp-name">隐藏列</div>
        <div class="cmp-note">ws.column_dimensions.group('D', 'F')<br>隐藏D-F列<br>修改字母即可</div>
      </div>
      <div class="cmp-card lof">
        <div class="cmp-icon">🔄</div>
        <div class="cmp-name">折叠/展开</div>
        <div class="cmp-note">group支持折叠展开<br>比简单隐藏更灵活<br>适合大量数据管理</div>
      </div>
    </div>
  </div>
</div>

<div class="section section-05">
  <div class="section-num num-05">05</div>
  <div class="section-body">
    <div class="tag tag-05">批量实战</div>
    <div class="section-title t-05">批量设置多行/列的行高列宽：遍历整表标准化</div>
    <div class="section-desc">插入/删除行列是一次性操作，但行高列宽通常需要整表统一设置。通过遍历行号或列号列表，在for循环中对每一行/列分别设置height/width，即可一次性完成整表的尺寸标准化。代码编写一次，永久适用于同类报表。</div>
    <div class="data-row">
      <div class="data-card">
        <div class="data-big">📏</div>
        <div class="data-name">批量行高</div>
        <div class="data-sm">for r in [1,2,3]:<br>ws.row_dimensions[r].height=40<br>遍历行号统一设行高</div>
      </div>
      <div class="data-card">
        <div class="data-big">📐</div>
        <div class="data-name">批量列宽</div>
        <div class="data-sm">for c in ['A','B','C']:<br>ws.column_dimensions[c].width=20<br>遍历列字母统一设列宽</div>
      </div>
      <div class="data-card">
        <div class="data-big">🔄</div>
        <div class="data-name">批量插入/删除</div>
        <div class="data-sm">一般一次性操作<br>无需遍历循环<br>直接指定位置和数量</div>
      </div>
    </div>
  </div>
</div>''',
    'takeaway': 'Excel的行列结构管理——插入、删除、调整行高列宽、隐藏——在Python中对应四对函数/属性：insert_rows/insert_cols、delete_rows/delete_cols、row_dimensions.height/column_dimensions.width、group组合隐藏。这些API直接映射Excel的右键菜单功能，学习曲线平缓。实际工作中，插入删除通常是一次性操作，而行高列宽则需要批量遍历整表设置。一个for循环遍历所有目标行列，统一设置height和width，让整个工作表的布局一步到位。Python让Excel的"体力活"变成了"智力活"。',
    'footer_ch': '用Python设置Excel单元格'
}

html7 = build_html(7, zh7)
path7 = os.path.join(BOOKDIR, f"{BASE}-ch007-info-zh.html")
with open(path7, 'w', encoding='utf-8') as f:
    f.write(html7)
print(f"✅ ch007 ZH: {len(html7)} bytes -> {path7}")


# ===== Chapter 8 =====
zh8 = {
    'ch_name': '用Python对Excel进行编辑',
    'scqa': 'Excel的「开始」选项卡「编辑」组提供排序、筛选、查找与替换三大核心功能。手工操作时，选中列→点击排序→选升序/降序，或点击筛选→勾选条件——每个Sheet重复一遍。当数据每天更新、每次都要重新排序筛选时，这些重复操作就是生产力杀手。如何用Python自动完成Excel的排序与筛选？本章讲解用Pandas的sort_values()实现单列与多列排序、用Pandas条件表达式实现数字筛选与文本筛选、以及查找替换的概要介绍。',
    'overview': '本章聚焦于Excel「开始」选项卡「编辑」组的Python实现。注意：openpyxl的排序和筛选只会添加排序/筛选符号，不会真的改变数据。真正的排序操作需使用Pandas库的sort_values(df, by=col, ascending=False)，支持单列排序和多列排序（列表传多列名，分别设置ascending）。筛选操作使用Pandas条件表达式（如df[df[\'col1\']>2]实现数字筛选，df[df[\'col2\']==\'b\']实现文本筛选），比Excel的鼠标点击更灵活。查找替换在第9章详细讲解。本章展示了Python生态"组合使用"的思想——openpyxl处理Excel文件读写，Pandas处理数据分析逻辑。',
    'kpis_html': '''
<div class="kpi-row">
  <div class="kpi-card">
    <div class="kpi-num">2种</div>
    <div class="kpi-label">排序方式<br>升序ascending=True<br>降序ascending=False</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-num">多列</div>
    <div class="kpi-label">多列排序<br>主关键字+次关键字<br>by=[col1,col2]</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-num">2种</div>
    <div class="kpi-label">筛选类型<br>数字筛选(>/<)<br>文本筛选(==/包含)</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-num">Pandas</div>
    <div class="kpi-label">核心库<br>sort_values + 条件表达式<br>替代Excel筛选菜单</div>
  </div>
</div>''',
    'sections_html': '''
<div class="section section-01">
  <div class="section-num num-01">01</div>
  <div class="section-body">
    <div class="tag tag-01">数据排序</div>
    <div class="section-title t-01">sort_values()：单列与多列排序一步到位</div>
    <div class="section-desc">openpyxl的排序操作只是添加排序符号，不会真正重排数据。真正的排序需用Pandas的sort_values()：df.sort_values(by='col1', ascending=False)按col1降序排列。多列排序传列表：by=['col1','col2']，先按col1排，重复值按col2排，可分别设升/降序。</div>
    <div class="code-block">import pandas as pd

# 单列排序：按col3降序
df_sorted = df.sort_values(by='col3', ascending=False)

# 多列排序：先col1降序，重复按col2升序
df_sorted = df.sort_values(
    by=['col1', 'col2'],
    ascending=[False, True])</div>
  </div>
</div>

<div class="section section-02">
  <div class="section-num num-02">02</div>
  <div class="section-body">
    <div class="tag tag-02">数据筛选</div>
    <div class="section-title t-02">Pandas条件筛选：数字与文本双模式</div>
    <div class="section-desc">openpyxl的筛选同样只添加筛选符号，不改变数据。Pandas直接用条件表达式实现筛选：df[df['col1']>2]筛选数字大于2的行，df[df['col2']=='b']筛选文本等于b的行。这与Excel中「数字筛选>大于」和「文本筛选>等于」完全对应，但更灵活——可自由组合条件。</div>
    <div class="dual-grid">
      <div class="dual-card yes">
        <div class="dual-icon">🔢</div>
        <div class="dual-text"><h4>数字筛选</h4><p>df[df['col1'] > 2]<br>大于 / 小于 / 等于<br>可组合多个条件<br>如 &gt;2 & &lt;10</p></div>
      </div>
      <div class="dual-card no">
        <div class="dual-icon">🔤</div>
        <div class="dual-text"><h4>文本筛选</h4><p>df[df['col2'] == 'b']<br>等于 / 包含 / 开头<br>支持字符串方法<br>.str.contains()</p></div>
      </div>
    </div>
  </div>
</div>

<div class="section section-03">
  <div class="section-num num-03">03</div>
  <div class="section-body">
    <div class="tag tag-03">排序流程</div>
    <div class="section-title t-03">从Excel到排序结果：完整操作链路</div>
    <div class="section-desc">完整的数据排序流程：①用openpyxl或pandas.read_excel()读取Excel数据到DataFrame；②用sort_values()按指定列排序（支持多列、升序/降序混合）；③将排序结果写回Excel或直接用于后续分析。这一链路将Excel的手工排序彻底自动化。</div>
    <div class="flow-row">
      <div class="flow-step">📂 读取<br><small>read_excel()</small></div>
      <div class="flow-arrow">→</div>
      <div class="flow-step">🔄 排序<br><small>sort_values()</small></div>
      <div class="flow-arrow">→</div>
      <div class="flow-step">📊 筛选<br><small>条件表达式</small></div>
      <div class="flow-arrow">→</div>
      <div class="flow-step">💾 写回<br><small>to_excel()</small></div>
      <div class="flow-arrow">→</div>
      <div class="flow-step end">✅ 自动化<br><small>一次编写永久复用</small></div>
    </div>
  </div>
</div>

<div class="section section-04">
  <div class="section-num num-04">04</div>
  <div class="section-body">
    <div class="tag tag-04">排序对比</div>
    <div class="section-title t-04">Excel vs Python 排序筛选对比</div>
    <div class="section-desc">Excel的排序筛选是手动交互式操作——每次数据更新要重新点一遍。Python的排序筛选是代码驱动——数据变化时只需要重新运行脚本。更重要的是，Pandas的条件筛选远强于Excel的筛选菜单：可组合任意复杂条件、可在排序前先筛选、可直接链式操作。</div>
    <div class="cmp-grid">
      <div class="cmp-card otc">
        <div class="cmp-icon">🖱️</div>
        <div class="cmp-name">Excel方式</div>
        <div class="cmp-note">选中列→点击排序<br>每次手动操作<br>多列需自定义排序</div>
      </div>
      <div class="cmp-card etf">
        <div class="cmp-icon">🐍</div>
        <div class="cmp-name">Python方式</div>
        <div class="cmp-note">df.sort_values()<br>代码一次编写<br>数据更新自动重算</div>
      </div>
      <div class="cmp-card lof">
        <div class="cmp-icon">🚀</div>
        <div class="cmp-name">Python优势</div>
        <div class="cmp-note">多列混合排序<br>排序+筛选链式操作<br>批量处理多个Sheet</div>
      </div>
    </div>
  </div>
</div>

<div class="section section-05">
  <div class="section-num num-05">05</div>
  <div class="section-body">
    <div class="tag tag-05">综合实战</div>
    <div class="section-title t-05">查找与替换 + 完整工作流</div>
    <div class="section-desc">「编辑」组还有查找和替换功能（第9章详讲），与排序筛选共同构成数据处理三步曲。实际工作流：读取Excel→筛选目标数据→排序→格式化→查找替换→写回。Python让这个完整流程从"手工一遍遍重复"变为"一键运行"。</div>
    <div class="data-row">
      <div class="data-card">
        <div class="data-big">🔍</div>
        <div class="data-name">查找</div>
        <div class="data-sm">定位特定数据<br>匹配单元格<br>第9章详讲</div>
      </div>
      <div class="data-card">
        <div class="data-big">✏️</div>
        <div class="data-name">替换</div>
        <div class="data-sm">批量修改内容<br>正则匹配替换<br>第9章详讲</div>
      </div>
      <div class="data-card">
        <div class="data-big">🔗</div>
        <div class="data-name">组合生态</div>
        <div class="data-sm">openpyxl + Pandas<br>读写 + 分析<br>Python报表自动化</div>
      </div>
    </div>
  </div>
</div>''',
    'takeaway': 'Excel的排序与筛选在Python中需要借助Pandas库来实现真正的数据重排和过滤。openpyxl负责Excel文件读写，Pandas负责数据分析逻辑——这是Python生态"组合使用"思想的典型体现。sort_values()支持单列和多列排序，条件表达式（df[条件]）支持数字和文本两种筛选模式，比Excel的鼠标操作更灵活、更强大。排序、筛选、格式化、查找替换串联成完整工作流，一次编写脚本，每天运行即可——这就是Python报表自动化的最终形态。',
    'footer_ch': '用Python对Excel进行编辑'
}

html8 = build_html(8, zh8)
path8 = os.path.join(BOOKDIR, f"{BASE}-ch008-info-zh.html")
with open(path8, 'w', encoding='utf-8') as f:
    f.write(html8)
print(f"✅ ch008 ZH: {len(html8)} bytes -> {path8}")

print("\n=== ALL ZH FILES WRITTEN ===")